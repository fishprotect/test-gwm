#
from selenium import webdriver
from openpyxl import Workbook
import re
class Selenium_PhantomJS:

    def __init__(self,url):
        self.url = url
        self.driver = webdriver.PhantomJS()
        self.driver.get(self.url)
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def row_title(self):
        detail = self.driver.find_elements_by_css_selector('.tbset > tbody:nth-child(1) > tr:nth-child(1) td .carbox.carbox-v2 a[target=_blank]')
        count = 2
        for i in detail:
            text = i.text 
            print(text)
            self.ws.cell(row=1,column=count,value=text)
            count += 1
    def column_title(self):
        detail_min = self.driver.find_elements_by_css_selector('[id^=tr_] > th:nth-child(1) > div:nth-child(1) > a:nth-child(1)')
        count_min = 2
        for i in detail_min:
            self.ws.cell(row=count_min,column=1,value=i.text)
            count_min += 1
    def content(self):
        detail = self.driver.find_elements_by_css_selector('table[id^=tab][class^=tbcs] [id^=tr_]:not([class^=headTr])')
        count_row = 2
        for i in detail:
            detail_min = i.find_elements_by_css_selector('td > div')
            count_column = 2
            for j in detail_min:
                s = str(j.text) 
                self.ws.cell(row=count_row,column=count_column,value=s)
                count_column += 1
            count_row += 1
    def fun_main(self):
        self.row_title()
        self.column_title()
        self.content()
        title = self.driver.title
        h =  re.findall('[H,M,F]\d\w{,1}\s*\w*',title)[0]
        self.wb.save('./haval1/'+h+'.xlsx')
        self.driver.quit()

if __name__ == '__main__':    
    url = 'http://price.pcauto.com.cn/sg4580/config.html'
    # url = 'https://price.pcauto.com.cn/sg11041/config.html'
    test = Selenium_PhantomJS(url)
    test.fun_main()
