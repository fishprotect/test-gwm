#
from selenium import webdriver
from openpyxl import Workbook
import re
class Selenium_PhantomJS:

    def __init__(self,url):
        self.url = url
        self.driver = webdriver.PhantomJS()
        self.driver.get(self.url)
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def row_title(self):
        detail = self.driver.find_elements_by_css_selector('#config_nav > table > tbody > tr td .carbox div a')
        car_type = {'hs_kw50':'标准型','hs_kw5':'豪华型','hs_kw0':'舒适型',}
        count = 2
        for i in detail:
            text = i.text 
            key = i.find_element_by_css_selector('span:last-child').get_attribute('class')[:6]
            key_1 = i.find_element_by_css_selector('span:last-child').value_of_css_property('content')  #获取tooltip的CSS属性color的属性值
            print(key,key_1)
            if key in car_type:
                key_value = car_type[key]
                text = text[:-1]+key_value+'型'
            print(text)
            self.ws.cell(row=1,column=count,value=text)
            count += 1
    # def column_title(self):
    #     detail_min = self.driver.find_elements_by_css_selector('#CarCompareContent > table > tbody > tr:not(.trForPic) th')
    #     count_min = 2
    #     for i in detail_min:
    #         print(i.text)
    #         self.ws.cell(row=count_min,column=1,value=i.text)
    #         count_min += 1
    # def content(self):
    #     detail = self.driver.find_elements_by_css_selector('#CarCompareContent > table > tbody > tr:not([id^=params])')
    #     count_row = 1
    #     for i in detail:
    #         detail_min = i.find_elements_by_css_selector('td[name^=td]')
    #         count_column = 2
    #         for j in detail_min:
    #             li = j.find_elements_by_css_selector('*')   
    #             if len(li)==0:
    #                 s = str(j.text)
    #                 self.ws.cell(row=count_row,column=count_column,value=s)
    #                 count_column += 1
    #             else:
    #                 val = ''
    #                 for detail_li in li:
    #                     content_li = str(detail_li.text)
    #                     val = val+content_li
    #                 self.ws.cell(row=count_row,column=count_column,value=val)
    #                 count_column += 1
    #         count_row += 1
    def fun_main(self):
        self.row_title()
        # self.column_title()
        # self.content()
        title = self.driver.title 
        self.wb.save(title+'.xlsx')

if __name__ == '__main__':    
    url = 'https://car.autohome.com.cn/config/series/3454.html#pvareaid=3454437'
    test = Selenium_PhantomJS(url)
    test.fun_main()
