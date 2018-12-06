#
from openpyxl import Workbook
from scrapy.selector import Selector
import requests
class Requests_crawl(object):

    def __init__(self,url,title):
        self.headers = {'User-Agent':'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:63.0) Gecko/20100101 Firefox/63.0'}
        self.wb = Workbook()
        self.ws = self.wb.active
        self.response = requests.get(url,headers=self.headers)
        self.contents = Selector(self.response)
        self.title = title
# 该函数用于填写汽车类型，比如：H1下面的舒适型，运动型
    def row_title(self):
        title = self.contents.css('tr#base_title > td[id^=mod_] >  a::text').extract()
        count_column = 2
        for i in title:
            self.ws.cell(row=1,column=count_column,value=i)
            count_column += 1

# 该函数用于挑出html结构中的所有的text，返回一个字符串
    def html_all_text(self,html_css):
        s = ''
        content = html_css.css('*')
        for each in content:
            text = each.css('::text').extract()
            if len(text)>0:
                for i in text:
                    s = s + i.strip()
        return s

# 该函数主要是完成表格的填写
    def content(self):
        contents = self.contents.css('.para_box#para_box>table[id^=base_]') # 16
        count_row = 2
        for each_type in contents:
            columns = each_type.css('tr')   #表示该参数类型下有多少行 比如：车身参数下面有多少行
            
            for column in columns:          #对具体的每一行进行操作
                title = self.html_all_text(column.css('td.title'))      # 该行的标题 比如：车身参数下的，车长
                self.ws.cell(row=count_row,column=1,value=title)

                content_text = column.css('td:not(.title)')     # 该行其他的单元格内容 比如：车长：3993mm
                count_column = 2  # 每一行均从第二列开始输入
                for content_column in content_text:
                    text = self.html_all_text(content_column)
                    text = str(text)
                    self.ws.cell(row=count_row,column=count_column,value=text)
                    count_column = count_column+1

                
                count_row += 1      # 结束该行的输入，行号加1
    def main(self):
        title = self.title 
        self.row_title()
        self.content()
        self.wb.save('./haval/'+title+'.xlsx')


if __name__ == "__main__":
    cars = {'H1':'http://newcar.xcar.com.cn/2658/config.htm'}
    for car in cars:
        url = cars[car]
        print(car,url)
        test = Requests_crawl(url,car)
        test.main()