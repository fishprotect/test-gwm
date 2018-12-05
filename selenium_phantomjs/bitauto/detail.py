#
from selenium import webdriver
from openpyxl import Workbook
class Selenium_PhantomJS:

    def __init__(self,url):
        self.url = url
        self.driver = webdriver.PhantomJS()
        self.driver.get(self.url)
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def row_title(self):
        detail = self.driver.find_elements_by_css_selector('#CarCompareContent > table > tbody > tr.trForPic td[id^=td] dd:nth-child(1)')
        count = 2
        for i in detail:
            print(i.text)
            self.ws.cell(row=1,column=count,value=i.text)
            count += 1
    def column_title(self):
        detail_min = self.driver.find_elements_by_css_selector('#CarCompareContent > table > tbody > tr:not(.trForPic) th')
        count_min = 2
        for i in detail_min:
            print(i.text)
            self.ws.cell(row=count_min,column=1,value=i.text)
            count_min += 1
    def content(self):
        detail = self.driver.find_elements_by_css_selector('#CarCompareContent > table > tbody > tr:not([id^=params])')
        count_row = 1
        for i in detail:
            detail_min = i.find_elements_by_css_selector('td[name^=td]')
            count_column = 2
            for j in detail_min:
                li = j.find_elements_by_css_selector('*')   
                if len(li)==0:
                    s = str(j.text)
                    self.ws.cell(row=count_row,column=count_column,value=s)
                    count_column += 1
                else:
                    val = ''
                    for detail_li in li:
                        content_li = str(detail_li.text)
                        val = val+content_li
                    self.ws.cell(row=count_row,column=count_column,value=val)
                    count_column += 1
            count_row += 1
    def fun_main(self):
        self.row_title()
        self.column_title()
        self.content()
        title = self.driver.title 
        self.wb.save(title+'.xlsx')

if __name__ == '__main__':    
    url = 'http://car.bitauto.com/hafuh1/peizhi/'
    test = Selenium_PhantomJS(url)
    test.fun_main()
